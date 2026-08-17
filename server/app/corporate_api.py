from __future__ import annotations

import io, json, os, re, secrets
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.auth import CurrentUser, get_current_user, hash_password
from app.db import db_session
from app.models import (
    User, EmployeeProfile, Chat, ChatMember, PayrollSlip,
    TechnicalDocument, OfflineExamSync, ExamQuestion, ExamAttempt, ExamAnswer,
)

router = APIRouter(prefix="/api/v1", tags=["corporate-v2"])
DOC_DIR = Path(os.getenv("TECH_DOC_DIR", "/app/techdocs"))


def is_admin(user: CurrentUser) -> bool:
    return user.role_key == "ADMIN" or "admin.manage" in user.perms


def get_profile(db: Session, uid: str) -> Optional[EmployeeProfile]:
    return db.scalar(select(EmployeeProfile).where(EmployeeProfile.user_id == uid))


def normalize_tab(v) -> str:
    s = str(v or "").strip()
    s = re.sub(r"\.0$", "", s)
    return s


def normalize_bool(v) -> bool:
    return str(v or "").strip().lower() in {"1","true","yes","да","ha","y"}


def map_role(position: str, explicit: str = "") -> str:
    if explicit:
        r=explicit.strip().upper()
        if r in {"WORKER","MASTER","ENGINEER","HEAD","ADMIN"}: return r
    p=(position or "").lower()
    if any(x in p for x in ["директор","director","начальник предприятия"]): return "HEAD"
    if any(x in p for x in ["начальник","завуч","заместитель","chief","head"]): return "HEAD"
    if any(x in p for x in ["инженер","engineer","итр","технолог","механик","энергетик"]): return "ENGINEER"
    if any(x in p for x in ["мастер","master"]): return "MASTER"
    return "WORKER"


def ensure_system_chat(db: Session, title: str, enterprise: str, workshop: str, *, director=False) -> Chat:
    dep = "DIRECTORS" if director else enterprise
    sec = "DIRECTORS" if director else workshop
    q=select(Chat).where(Chat.title==title, Chat.department_id==dep, Chat.section_id==sec)
    c=db.scalar(q)
    if not c:
        c=Chat(title=title, department_id=dep, section_id=sec, created_by="SYSTEM", is_super_admin=False, is_announcement=False)
        db.add(c); db.flush()
    return c


def sync_memberships_for_employee(db: Session, user: User, p: EmployeeProfile):
    if p.employment_status != "ACTIVE": return
    enterprise_chat=ensure_system_chat(db, f"Общий чат {p.enterprise_id}", p.enterprise_id, "ALL")
    workshop_chat=ensure_system_chat(db, f"Общий чат {p.workshop_id}", p.enterprise_id, p.workshop_id)
    for c in (enterprise_chat, workshop_chat):
        if not db.scalar(select(ChatMember).where(ChatMember.chat_id==c.id, ChatMember.user_id==user.id)):
            db.add(ChatMember(chat_id=c.id,user_id=user.id,is_admin=user.role_key in {"HEAD","ADMIN"}))
    if p.is_director:
        dc=ensure_system_chat(db, "🔒 Чат директоров АГМК", "AGMK", "DIRECTORS", director=True)
        if not db.scalar(select(ChatMember).where(ChatMember.chat_id==dc.id, ChatMember.user_id==user.id)):
            db.add(ChatMember(chat_id=dc.id,user_id=user.id,is_admin=True))


class ImportResult(BaseModel):
    created: int
    updated: int
    skipped: int
    temporary_passwords: dict[str,str]


@router.post("/admin/employees/import-xlsx", response_model=ImportResult)
async def import_employees(file: UploadFile=File(...), db: Session=Depends(db_session), user: CurrentUser=Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403,"Admin permission required")
    data=await file.read()
    try:
        wb=load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400,f"Invalid Excel: {e}")
    ws=wb.active
    rows=list(ws.iter_rows(values_only=True))
    if not rows: raise HTTPException(400,"Empty Excel")
    headers=[str(x or "").strip().lower() for x in rows[0]]
    aliases={
        "tab_no":["табельный номер","табельный №","табельный","tab_no","tabel","employee_id"],
        "full_name":["фио","ф.и.о.","ф.и.о","full_name","name"],
        "enterprise":["предприятие","enterprise","company"],
        "workshop":["цех","подразделение","section","workshop","department"],
        "position":["должность","position","job"],
        "manager":["руководитель","manager_tab_no","табельный руководителя"],
        "status":["статус","status"],
        "domain_login":["домен","domain_login","domain"],
        "domain_allowed":["домен доступ","domain_apps_allowed","domain_allowed"],
        "director":["директор","is_director","director"],
        "role":["роль","role","role_key"],
    }
    def idx(key):
        for a in aliases[key]:
            if a in headers: return headers.index(a)
        return None
    required={k:idx(k) for k in ["tab_no","full_name"]}
    if None in required.values(): raise HTTPException(400,"Excel must contain Ф.И.О. and табельный номер")
    ids={k:idx(k) for k in aliases}
    created=updated=skipped=0; passwords={}
    for r in rows[1:]:
        tab=normalize_tab(r[ids["tab_no"]])
        name=str(r[ids["full_name"]] or "").strip()
        if not tab or not name: skipped+=1; continue
        uid=tab
        ent=str(r[ids["enterprise"]] or "AGMK").strip() if ids["enterprise"] is not None else "AGMK"
        workshop=str(r[ids["workshop"]] or "GENERAL").strip() if ids["workshop"] is not None else "GENERAL"
        pos=str(r[ids["position"]] or "").strip() if ids["position"] is not None else ""
        explicit=str(r[ids["role"]] or "").strip() if ids["role"] is not None else ""
        role=map_role(pos, explicit)
        u=db.get(User,uid)
        if not u:
            temp=secrets.token_urlsafe(6)
            u=User(id=uid,full_name=name,role_key=role,department_id=ent,section_id=workshop,password_hash=hash_password(temp))
            db.add(u); db.flush(); created+=1; passwords[tab]=temp
        else:
            u.full_name=name; u.role_key=role; u.department_id=ent; u.section_id=workshop; updated+=1
        p=get_profile(db,uid)
        if not p:
            p=EmployeeProfile(user_id=uid,tab_no=tab); db.add(p)
        p.enterprise_id=ent; p.workshop_id=workshop; p.position=pos
        p.manager_tab_no=str(r[ids["manager"]] or "").strip() if ids["manager"] is not None else ""
        p.employment_status=str(r[ids["status"]] or "ACTIVE").strip().upper() if ids["status"] is not None else "ACTIVE"
        p.domain_login=str(r[ids["domain_login"]] or "").strip() if ids["domain_login"] is not None else ""
        p.domain_apps_allowed=normalize_bool(r[ids["domain_allowed"]]) if ids["domain_allowed"] is not None else bool(p.domain_login)
        p.is_director=normalize_bool(r[ids["director"]]) if ids["director"] is not None else ("директор" in pos.lower())
        p.updated_at=datetime.utcnow()
        sync_memberships_for_employee(db,u,p)
    db.commit()
    return ImportResult(created=created,updated=updated,skipped=skipped,temporary_passwords=passwords)


@router.get("/me/corporate")
def my_corporate_profile(db: Session=Depends(db_session), user: CurrentUser=Depends(get_current_user)):
    u=db.get(User,user.id); p=get_profile(db,user.id)
    return {"id":u.id,"tab_no":p.tab_no if p else u.id,"full_name":u.full_name,"role":u.role_key,
            "enterprise":p.enterprise_id if p else u.department_id,"workshop":p.workshop_id if p else u.section_id,
            "position":p.position if p else "","domain_apps_allowed":bool(p.domain_apps_allowed) if p else False,
            "is_director":bool(p.is_director) if p else False}




class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str = Field(min_length=6,max_length=128)

@router.post("/me/change-password")
def change_password(payload:ChangePasswordIn,db:Session=Depends(db_session),user:CurrentUser=Depends(get_current_user)):
    from app.auth import verify_password
    u=db.get(User,user.id)
    if not u or not u.password_hash or not verify_password(payload.current_password,u.password_hash):
        raise HTTPException(400,"Current password is incorrect")
    u.password_hash=hash_password(payload.new_password)
    p=get_profile(db,user.id)
    if p: p.must_change_password=False; p.updated_at=datetime.utcnow()
    db.commit(); return {"ok":True}

@router.get("/modules")
def modules(db: Session=Depends(db_session), user: CurrentUser=Depends(get_current_user)):
    p=get_profile(db,user.id); domain=bool(p and p.domain_apps_allowed)
    itr=user.role_key in {"MASTER","ENGINEER","HEAD","ADMIN"}
    return [
        {"key":"chats","title":"Чаты","enabled":True,"visible":True},
        {"key":"tickets","title":"Заявки","enabled":True,"visible":True},
        {"key":"incidents","title":"Инциденты","enabled":True,"visible":True},
        {"key":"exams","title":"Экзамены Online / Offline","enabled":True,"visible":True},
        {"key":"itr_ai","title":"AI-помощник ИТР","enabled":True,"visible":itr},
        {"key":"payroll","title":"Мои расчётные листы","enabled":True,"visible":True},
        {"key":"1c","title":"1С","enabled":False,"visible":domain,"status":"Ожидает официальной интеграции"},
        {"key":"directum","title":"Directum","enabled":False,"visible":domain,"status":"Ожидает официальной интеграции"},
    ]


class AiChatBuildIn(BaseModel):
    title: str = Field(min_length=1,max_length=255)
    enterprise_id: Optional[str]=None
    workshop_id: Optional[str]=None
    role_key: Optional[str]=None
    user_ids: list[str]=Field(default_factory=list)

@router.post("/ai/chat-builder")
def ai_chat_builder(payload: AiChatBuildIn, db: Session=Depends(db_session), user: CurrentUser=Depends(get_current_user)):
    # Deterministic corporate template engine now; an LLM can be connected later without changing the permission layer.
    me=get_profile(db,user.id)
    if not me: raise HTTPException(400,"Employee profile required")
    ent=(payload.enterprise_id or me.enterprise_id).strip(); workshop=(payload.workshop_id or me.workshop_id).strip()
    if user.role_key not in {"HEAD","ADMIN"} and (ent!=me.enterprise_id or workshop!=me.workshop_id):
        raise HTTPException(403,"Cannot create chat outside your enterprise/workshop")
    q=select(User).where(User.department_id==ent,User.section_id==workshop)
    if payload.role_key: q=q.where(User.role_key==payload.role_key.upper())
    members=set(db.scalars(q).all())
    ids={u.id for u in members} | set(payload.user_ids) | {user.id}
    # Block cross-org explicit users for non-head/admin.
    final=[]
    for uid in ids:
        u=db.get(User,uid)
        if not u: continue
        if user.role_key not in {"HEAD","ADMIN"} and (u.department_id!=ent or u.section_id!=workshop): continue
        final.append(uid)
    c=Chat(title=payload.title,department_id="PRIVATE",section_id="GROUP",created_by=user.id)
    db.add(c); db.flush()
    for uid in sorted(set(final)): db.add(ChatMember(chat_id=c.id,user_id=uid,is_admin=(uid==user.id)))
    db.commit(); return {"id":c.id,"title":c.title,"members":len(set(final))}


@router.post("/admin/payroll/import-xlsx")
async def payroll_import(period: str=Form(...), file: UploadFile=File(...), db: Session=Depends(db_session), user: CurrentUser=Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403,"Admin permission required")
    data=await file.read(); wb=load_workbook(io.BytesIO(data),read_only=True,data_only=True); ws=wb.active
    rows=list(ws.iter_rows(values_only=True)); headers=[str(x or "").strip().lower() for x in rows[0]]
    def col(names):
        for n in names:
            if n in headers:return headers.index(n)
        return None
    itab=col(["табельный номер","табельный","tab_no"]); igross=col(["начислено","gross","gross_amount"]); ided=col(["удержано","deductions"]); inet=col(["к выплате","net","net_amount"])
    if itab is None: raise HTTPException(400,"No табельный номер column")
    count=0
    for r in rows[1:]:
        tab=normalize_tab(r[itab]); u=db.get(User,tab)
        if not tab or not u: continue
        slip=db.scalar(select(PayrollSlip).where(PayrollSlip.tab_no==tab,PayrollSlip.period==period))
        details={headers[i]:str(v or "") for i,v in enumerate(r)}
        if not slip:
            slip=PayrollSlip(user_id=u.id,tab_no=tab,period=period); db.add(slip)
        slip.gross_amount=str(r[igross] or "") if igross is not None else ""
        slip.deductions=str(r[ided] or "") if ided is not None else ""
        slip.net_amount=str(r[inet] or "") if inet is not None else ""
        slip.details_json=json.dumps(details,ensure_ascii=False); slip.source_file=file.filename or ""
        # Private notification chat for this employee; no salary amount is exposed in the chat text.
        pc=db.scalar(select(Chat).where(Chat.title=="💰 Расчётные листы",Chat.department_id=="PRIVATE",Chat.section_id==f"PAYROLL:{u.id}"))
        if not pc:
            pc=Chat(title="💰 Расчётные листы",department_id="PRIVATE",section_id=f"PAYROLL:{u.id}",created_by="SYSTEM",is_read_only=1)
            db.add(pc); db.flush(); db.add(ChatMember(chat_id=pc.id,user_id=u.id,is_admin=False))
        db.add(__import__("app.models",fromlist=["Message"]).Message(chat_id=pc.id,user_id="SYSTEM",text=f"Доступен расчётный лист за {period}. Откройте раздел «Мои расчётные листы»."))
        count+=1
    db.commit(); return {"ok":True,"imported":count,"period":period}

@router.get("/me/payroll")
def my_payroll(db: Session=Depends(db_session), user: CurrentUser=Depends(get_current_user)):
    rows=db.scalars(select(PayrollSlip).where(PayrollSlip.user_id==user.id).order_by(PayrollSlip.period.desc())).all()
    return [{"period":x.period,"gross":x.gross_amount,"deductions":x.deductions,"net":x.net_amount,"details":json.loads(x.details_json or "{}") } for x in rows]


@router.post("/admin/tech-docs")
async def upload_tech_doc(enterprise_id:str=Form(...),workshop_id:str=Form("ALL"),equipment_model:str=Form(""),equipment_type:str=Form(""),title:str=Form(...),doc_type:str=Form("MANUAL"),text_content:str=Form(""),file:UploadFile|None=File(None),db:Session=Depends(db_session),user:CurrentUser=Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403,"Admin permission required")
    path=""
    if file:
        DOC_DIR.mkdir(parents=True,exist_ok=True); safe=re.sub(r"[^A-Za-z0-9._-]+","_",Path(file.filename or "doc").name); name=f"{secrets.token_hex(8)}_{safe}"; (DOC_DIR/name).write_bytes(await file.read()); path=name
    d=TechnicalDocument(enterprise_id=enterprise_id,workshop_id=workshop_id,equipment_model=equipment_model,equipment_type=equipment_type,title=title,doc_type=doc_type,file_path=path,text_content=text_content)
    db.add(d); db.commit(); db.refresh(d); return {"id":d.id,"title":d.title}

class TechAskIn(BaseModel):
    query: str
    equipment_model: str=""

@router.post("/itr/assistant")
def itr_assistant(payload:TechAskIn,db:Session=Depends(db_session),user:CurrentUser=Depends(get_current_user)):
    if user.role_key not in {"MASTER","ENGINEER","HEAD","ADMIN"}: raise HTTPException(403,"ITR role required")
    p=get_profile(db,user.id); ent=p.enterprise_id if p else user.department_id; workshop=p.workshop_id if p else user.section_id
    q=select(TechnicalDocument).where(TechnicalDocument.enterprise_id==ent)
    docs=db.scalars(q).all(); terms=[x.lower() for x in re.findall(r"[\w.-]{3,}",payload.query)]
    scored=[]
    for d in docs:
        if d.workshop_id not in {"ALL",workshop}: continue
        hay=(d.title+" "+d.equipment_model+" "+d.equipment_type+" "+d.text_content).lower(); score=sum(hay.count(t) for t in terms)
        if payload.equipment_model and payload.equipment_model.lower() in hay: score+=5
        if score: scored.append((score,d))
    scored.sort(key=lambda x:x[0],reverse=True); top=[d for _,d in scored[:5]]
    return {"answer":"Найдены документы по запросу. Для точного AI-RAG ответа подключается локальная модель/векторный поиск на следующем этапе.","sources":[{"id":d.id,"title":d.title,"equipment_model":d.equipment_model,"doc_type":d.doc_type} for d in top],"needs_ai_model":True}


@router.get("/exams/{exam_key}/offline-package")
def offline_exam_package(exam_key:str,limit:int=30,db:Session=Depends(db_session),user:CurrentUser=Depends(get_current_user)):
    rows=db.scalars(select(ExamQuestion).where(ExamQuestion.exam_key==exam_key.upper(),ExamQuestion.is_active==1).limit(min(max(limit,1),100))).all()
    return {"exam_key":exam_key.upper(),"generated_at":datetime.utcnow().isoformat(),"questions":[{"id":q.id,"qcode":q.qcode,"topic":q.topic,"question":q.question,"qtype":q.qtype,"options":q.options,"media_path":q.media_path} for q in rows]}

class OfflineSyncIn(BaseModel):
    client_attempt_id:str
    exam_key:str
    answers:list[dict]

@router.post("/exams/offline-sync")
def offline_exam_sync(payload:OfflineSyncIn,db:Session=Depends(db_session),user:CurrentUser=Depends(get_current_user)):
    existing=db.scalar(select(OfflineExamSync).where(OfflineExamSync.client_attempt_id==payload.client_attempt_id))
    if existing:return {"ok":True,"duplicate":True}
    attempt=ExamAttempt(exam_key=payload.exam_key.upper(),user_id=user.id,started_at=datetime.utcnow().isoformat(),finished_at=datetime.utcnow().isoformat()); db.add(attempt); db.flush()
    correct=0
    for item in payload.answers:
        q=db.get(ExamQuestion,int(item.get("question_id",0))); sel=str(item.get("selected","")).strip().upper()
        if not q or q.exam_key!=attempt.exam_key: continue
        is_ok=1 if sel==str(q.correct).strip().upper() else 0; correct+=is_ok
        db.add(ExamAnswer(attempt_id=attempt.id,question_id=q.id,selected=sel,is_correct=is_ok,answered_at=datetime.utcnow().isoformat()))
    db.add(OfflineExamSync(client_attempt_id=payload.client_attempt_id,user_id=user.id,exam_key=attempt.exam_key,payload_json=json.dumps(payload.model_dump(),ensure_ascii=False)))
    db.commit(); return {"ok":True,"duplicate":False,"attempt_id":attempt.id,"correct":correct,"total":len(payload.answers)}
